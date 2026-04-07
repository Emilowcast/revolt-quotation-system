#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
generate-sales-report-COMPLETE.py
Genera reporte de ventas en Excel con formato EXACTO al original
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============================================
# CONSTANTES DE ESTILO
# ============================================

# Colores REVOLT (escala de rojos)
COLOR_TITULO = "F62E41"  # Rojo REVOLT principal
COLOR_HEADER = "D41F33"  # Rojo oscuro
COLOR_TOTAL = "FFE6E6"   # Rosa claro
COLOR_MES = "FFC7CE"     # Rosa medio

# Fuentes
FONT_TITULO = Font(name='Calibri', size=14, bold=True, color='FFFFFF')  # Blanco
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # Blanco
FONT_NORMAL = Font(name='Calibri', size=10)
FONT_TOTAL = Font(name='Calibri', size=10, bold=True)

# Bordes
BORDER_THIN = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# ============================================
# FUNCIÓN PRINCIPAL
# ============================================

def generar_reporte(data_json, output_path, año):
    """
    Genera el Excel con formato exacto - REPORTE ANUAL
    
    Args:
        data_json: str - JSON con datos de ventas
        output_path: str - Ruta donde guardar el Excel
        año: int - Año del reporte
    """
    print(f"[REPORTE] Generando reporte Excel ANUAL: {año}")
    
    # Cargar datos
    try:
        data = json.loads(data_json)
    except:
        print("[ERROR] JSON invalido")
        sys.exit(1)
    
    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto
    
    # Crear las 2 hojas principales
    print("  [GENERAL] Creando hoja GENERAL...")
    crear_hoja_general(wb, data['general'], año)
    
    print("  [COMISIONES] Creando hoja COMISIONES...")
    crear_hoja_comisiones(wb, data['comisiones'], año)

    # ⭐ NUEVO: Sincronizar ancho columna A
    ws_general = wb["GENERAL"]
    ws_comisiones = wb["Comisiones"]
    ws_comisiones.column_dimensions['A'].width = ws_general.column_dimensions['A'].width

    wb.save(output_path)
    print(f"[OK] Excel generado: {output_path}")
    
    print("  [COMERCIALIZACION] Creando hoja COMERCIALIZACION...")
    crear_hoja_comercializacion(wb, data['comercializacion'], año)
    
    # Guardar
    wb.save(output_path)
    print(f"[OK] Excel generado: {output_path}")

# ============================================
# FUNCIONES AUXILIARES DE FORMATO
# ============================================

def formatear_cliente(venta):
    """
    Formatea el cliente como: Nombre Cliente - Empresa
    """
    nombre = venta.get('cliente', 'Cliente')
    empresa = venta.get('empresa', '')
    
    if empresa and empresa.strip():
        return f"{nombre} - {empresa}"
    return nombre


def extraer_info_producto(concepto):
    """
    Extrae información estructurada del concepto y la formatea sin etiquetas
    Ejemplo entrada: "Modelo: RM-041-120 Capacidad: 4 kVA Sistema: Monofásico Voltaje de Salida: 120 V"
    Ejemplo salida: "RM-041-120 - 4 kVA - Monofásico - 120 V"
    """
    import re
    
    # Si el concepto ya está limpio (no tiene "Modelo:", etc), devolverlo tal cual
    if 'Modelo:' not in concepto and 'Capacidad:' not in concepto:
        return concepto
    
    # Extraer valores usando regex
    modelo = re.search(r'Modelo:\s*([^\n]+?)(?:\s*Capacidad:|$)', concepto)
    capacidad = re.search(r'Capacidad:\s*([^\n]+?)(?:\s*Sistema:|$)', concepto)
    sistema = re.search(r'Sistema:\s*([^\n]+?)(?:\s*Voltaje|$)', concepto)
    voltaje = re.search(r'Voltaje[^:]*:\s*([^\n]+?)$', concepto)
    
    # Construir el resultado con solo los valores
    partes = []
    if modelo:
        partes.append(modelo.group(1).strip())
    if capacidad:
        partes.append(capacidad.group(1).strip())
    if sistema:
        partes.append(sistema.group(1).strip())
    if voltaje:
        partes.append(voltaje.group(1).strip())
    
    # Unir con " - "
    if partes:
        return " - ".join(partes)
    
    # Si no se encontró nada, devolver el concepto original
    return concepto


def ajustar_altura_fila(ws, row_idx):
    """
    Ajusta la altura de la fila para que todo el contenido sea visible
    """
    max_lines = 1
    for cell in ws[row_idx]:
        if cell.value:
            # Contar líneas de texto
            lines = str(cell.value).count('\n') + 1
            # Estimar líneas por ancho de columna
            col_width = ws.column_dimensions[cell.column_letter].width or 10
            text_length = len(str(cell.value))
            estimated_lines = max(1, text_length // int(col_width * 1.5))
            max_lines = max(max_lines, lines, estimated_lines)
    
    # Establecer altura (15 puntos por línea aproximadamente)
    ws.row_dimensions[row_idx].height = max(15, max_lines * 15)


def ajustar_anchos_columnas(ws):
    """
    Ajusta el ancho de las columnas según el contenido
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                # Calcular longitud del contenido
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # Ajustar ancho (con un mínimo y máximo)
        adjusted_width = min(max(max_length + 2, 10), 60)
        ws.column_dimensions[column_letter].width = adjusted_width


def formatear_semana(fecha_viernes, mes):
    """
    Formatea la semana como: Semana 1 (03-ene)
    """
    from datetime import datetime
    
    # Convertir fecha string a objeto datetime
    if isinstance(fecha_viernes, str):
        # Formato: dd/mm/yyyy
        partes = fecha_viernes.split('/')
        fecha_obj = datetime(int(partes[2]), int(partes[1]), int(partes[0]))
    else:
        fecha_obj = fecha_viernes
    
    # Calcular número de semana del mes
    dia_mes = fecha_obj.day
    numero_semana = ((dia_mes - 1) // 7) + 1
    
    # Obtener nombre corto del mes
    meses_cortos = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 
                    'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
    mes_corto = meses_cortos[fecha_obj.month - 1]
    
    # Formatear: Semana 1 (03-ene)
    return f"Semana {numero_semana} ({fecha_obj.day:02d}-{mes_corto})"

# ============================================
# HOJA 1: GENERAL
# ============================================

def crear_hoja_general(wb, ventas, año):
    """Crea la hoja GENERAL con formato exacto - ANUAL CON TODOS LOS MESES"""
    ws = wb.create_sheet("GENERAL")
    
    # Configurar anchos de columna (iniciales, se ajustarán después)
    ws.column_dimensions['A'].width = 12  # Fecha
    ws.column_dimensions['B'].width = 35  # Cliente
    ws.column_dimensions['C'].width = 50  # Concepto
    ws.column_dimensions['D'].width = 18  # Con IVA (ampliado para "VENTAS ENERO")
    ws.column_dimensions['E'].width = 14  # Sin IVA
    ws.column_dimensions['F'].width = 14  # Total
    ws.column_dimensions['G'].width = 18  # Semana al
    ws.column_dimensions['H'].width = 15  # Vendedor
    ws.column_dimensions['I'].width = 3   # Espacio
    ws.column_dimensions['J'].width = 15  # VENDEDOR (tabla)
    ws.column_dimensions['K'].width = 15  # VENTAS (tabla)
    ws.column_dimensions['L'].width = 10  # % (tabla)
    
    # Fila 2: Título principal
    ws.merge_cells('A2:H2')
    titulo = ws['A2']
    titulo.value = f"Ingreso de Ventas Anual {año}"
    titulo.font = FONT_TITULO
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    titulo.fill = PatternFill(start_color=COLOR_TITULO, end_color=COLOR_TITULO, fill_type='solid')
    
    # Filas 5-6: Encabezados en 2 niveles
    # ⭐ COMBINAR VERTICALMENTE encabezados que ocupan 2 filas
    
    # Fecha (A5:A6)
    ws.merge_cells('A5:A6')
    ws.cell(5, 1).value = 'Fecha'
    ws.cell(5, 1).font = FONT_HEADER
    ws.cell(5, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 1).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(5, 1).border = BORDER_THIN
    ws.cell(6, 1).border = BORDER_THIN  # ⭐ BORDE TAMBIÉN EN FILA 6
    
    # Cliente (B5:B6)
    ws.merge_cells('B5:B6')
    ws.cell(5, 2).value = 'Cliente'
    ws.cell(5, 2).font = FONT_HEADER
    ws.cell(5, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 2).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(5, 2).border = BORDER_THIN
    ws.cell(6, 2).border = BORDER_THIN  # ⭐ BORDE TAMBIÉN EN FILA 6
    
    # Concepto (C5:C6)
    ws.merge_cells('C5:C6')
    ws.cell(5, 3).value = 'Concepto'
    ws.cell(5, 3).font = FONT_HEADER
    ws.cell(5, 3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 3).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(5, 3).border = BORDER_THIN
    ws.cell(6, 3).border = BORDER_THIN  # ⭐ BORDE TAMBIÉN EN FILA 6
    
    # ⭐ DEPÓSITO (D5:E5) - Solo fila 5
    ws.merge_cells('D5:E5')
    ws.cell(5, 4).value = 'DEPÓSITO'
    ws.cell(5, 4).font = FONT_HEADER
    ws.cell(5, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 4).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(5, 4).border = BORDER_THIN
    ws.cell(5, 5).border = BORDER_THIN  # ⭐ BORDE EN E5 TAMBIÉN
    
    # Total (F5:F6)
    ws.merge_cells('F5:F6')
    ws.cell(5, 6).value = 'Total'
    ws.cell(5, 6).font = FONT_HEADER
    ws.cell(5, 6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 6).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(5, 6).border = BORDER_THIN
    ws.cell(6, 6).border = BORDER_THIN  # ⭐ BORDE TAMBIÉN EN FILA 6
    
    # Semana (G5:G6) ⭐ CAMBIAR DE "Semana al" A "Semana"
    ws.merge_cells('G5:G6')
    ws.cell(5, 7).value = 'Semana'
    ws.cell(5, 7).font = FONT_HEADER
    ws.cell(5, 7).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 7).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(5, 7).border = BORDER_THIN
    ws.cell(6, 7).border = BORDER_THIN  # ⭐ BORDE TAMBIÉN EN FILA 6
    
    # Vendedor (H5:H6)
    ws.merge_cells('H5:H6')
    ws.cell(5, 8).value = 'Vendedor'
    ws.cell(5, 8).font = FONT_HEADER
    ws.cell(5, 8).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 8).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(5, 8).border = BORDER_THIN
    ws.cell(6, 8).border = BORDER_THIN  # ⭐ BORDE TAMBIÉN EN FILA 6
    
    # Fila 6: Sub-encabezados (Con IVA, Sin IVA)
    ws.cell(6, 4).value = 'Con IVA'
    ws.cell(6, 4).font = FONT_HEADER
    ws.cell(6, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(6, 4).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(6, 4).border = BORDER_THIN
    
    ws.cell(6, 5).value = 'Sin IVA'
    ws.cell(6, 5).font = FONT_HEADER
    ws.cell(6, 5).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(6, 5).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    ws.cell(6, 5).border = BORDER_THIN
    
    # ⭐ AGRUPAR VENTAS POR MES
    ventas_por_mes = {}
    for venta in ventas:
        mes = venta.get('mes', 1)
        if mes not in ventas_por_mes:
            ventas_por_mes[mes] = []
        ventas_por_mes[mes].append(venta)
    
    # Nombres de meses
    mes_nombres = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
                   'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    
    current_row = 7
    
    # ⭐ PROCESAR CADA MES
    for mes in range(1, 13):
        if mes not in ventas_por_mes:
            continue  # Saltar meses sin ventas
        
        ventas_del_mes = ventas_por_mes[mes]
        
        # ⭐ Título del mes (MEJORADO)
        ws.merge_cells(f'A{current_row}:H{current_row}')  # Combinar A-H
        mes_cell = ws.cell(current_row, 1)
        mes_cell.value = mes_nombres[mes]
        mes_cell.font = Font(name='Calibri', size=20, bold=True, color='FFFFFF')  # ⭐ Tamaño 20, blanco
        mes_cell.alignment = Alignment(horizontal='center', vertical='center')  # ⭐ Centrado
        mes_cell.fill = PatternFill(start_color='F62E41', end_color='F62E41', fill_type='solid')  # ⭐ Rojo REVOLT
        ws.row_dimensions[current_row].height = 30  # ⭐ Altura para que se vea bien
        current_row += 1
        # Datos de ventas del mes
        current_week = None
        week_total = 0
        month_total = 0
        ventas_por_vendedor = {}
        row_inicio_mes = current_row
        week_start_row = None  # ⭐ Rastrear inicio de semana para merge
        filas_totales_semanales = []  # ⭐ NUEVO: Rastrear filas de totales para fórmula mensual
        
        for venta in ventas_del_mes:
            # Detectar cambio de semana
            if current_week and venta['week'] != current_week:
                # ⭐ COMBINAR CELDAS DE LA SEMANA ANTERIOR
                if week_start_row and current_row > week_start_row:
                    # Combinar columna Semana (G)
                    ws.merge_cells(f'G{week_start_row}:G{current_row - 1}')
                    ws.cell(week_start_row, 7).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # ⭐ COMBINAR columna Total (F) con FÓRMULA SUM
                    ws.merge_cells(f'F{week_start_row}:F{current_row - 1}')
                    ws.cell(week_start_row, 6).value = f"=SUM(E{week_start_row}:E{current_row - 1})"  # ⭐ FÓRMULA
                    ws.cell(week_start_row, 6).number_format = '$#,##0.00'
                    ws.cell(week_start_row, 6).font = FONT_TOTAL
                    ws.cell(week_start_row, 6).alignment = Alignment(horizontal='center', vertical='center')
                
                current_row = agregar_total_semanal_general(ws, current_row, week_start_row, filas_totales_semanales)  # ⭐ Pasar inicio de semana
                week_total = 0
                week_start_row = None
            
            # Iniciar nueva semana
            if current_week != venta['week']:
                current_week = venta['week']
                week_start_row = current_row  # ⭐ GUARDAR INICIO DE SEMANA
            
            # Agregar venta
            ws.cell(current_row, 1).value = venta['fecha']
            
            ws.cell(current_row, 2).value = formatear_cliente(venta)
            ws.cell(current_row, 2).alignment = Alignment(wrap_text=True, vertical='top')
            
            concepto_formateado = extraer_info_producto(venta['concepto'])
            ws.cell(current_row, 3).value = concepto_formateado
            ws.cell(current_row, 3).alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.cell(current_row, 4).value = venta['amountWithIVA']
            ws.cell(current_row, 4).number_format = '$#,##0.00'
            
            ws.cell(current_row, 5).value = f"=D{current_row}/1.16"
            ws.cell(current_row, 5).number_format = '$#,##0.00'
            
            # ⭐ COLUMNA F (Total): NO poner valor aquí, se pondrá al combinar
            # ws.cell(current_row, 6) se llenará al combinar las celdas
            
            # ⭐ SEMANA: Solo poner valor en la primera celda
            if current_row == week_start_row:
                ws.cell(current_row, 7).value = formatear_semana(venta['week'], mes)
            
            vendedor = venta['vendor']
            ws.cell(current_row, 8).value = vendedor
            
            ajustar_altura_fila(ws, current_row)
            # ⭐ Colorear fila si es venta extranjera
            if venta.get('country', 'MX') != 'MX':
                colorear_fila_extranjero(ws, current_row, 8, BORDER_THIN)
            
            amount_sin_iva = venta['amountWithoutIVA']
            week_total += amount_sin_iva
            month_total += amount_sin_iva
            
            if vendedor not in ventas_por_vendedor:
                ventas_por_vendedor[vendedor] = 0
            ventas_por_vendedor[vendedor] += amount_sin_iva
            
            current_row += 1
        
        # ⭐ COMBINAR ÚLTIMA SEMANA DEL MES
        if week_start_row and current_row > week_start_row:
            # Combinar Semana
            ws.merge_cells(f'G{week_start_row}:G{current_row - 1}')
            ws.cell(week_start_row, 7).alignment = Alignment(horizontal='center', vertical='center')
            
            # ⭐ Combinar Total con FÓRMULA
            ws.merge_cells(f'F{week_start_row}:F{current_row - 1}')
            ws.cell(week_start_row, 6).value = f"=SUM(E{week_start_row}:E{current_row - 1})"  # ⭐ FÓRMULA
            ws.cell(week_start_row, 6).number_format = '$#,##0.00'
            ws.cell(week_start_row, 6).font = FONT_TOTAL
            ws.cell(week_start_row, 6).alignment = Alignment(horizontal='center', vertical='center')
        
        # Total de la última semana del mes con FÓRMULA
        if week_total > 0:
            current_row = agregar_total_semanal_general(ws, current_row, week_start_row, filas_totales_semanales)
        
        # ⭐ TOTAL DEL MES con FÓRMULA
        ws.cell(current_row, 4).value = f"VENTAS {mes_nombres[mes]}"
        ws.cell(current_row, 4).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws.cell(current_row, 4).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(current_row, 4).fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
        
        # ⭐ FÓRMULA: Sumar totales semanales
        if filas_totales_semanales:
            celdas_sum = "+".join([f"E{fila}" for fila in filas_totales_semanales])
            ws.cell(current_row, 5).value = f"={celdas_sum}"
        else:
            ws.cell(current_row, 5).value = 0
        
        ws.cell(current_row, 5).number_format = '$#,##0.00'
        ws.cell(current_row, 5).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws.cell(current_row, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(current_row, 5).fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
        
        fila_total_mes = current_row  # ⭐ Guardar fila para referencia en tabla
        current_row += 1
        
        # ⭐ TABLA RESUMEN POR VENDEDOR (al lado derecho) con FÓRMULAS
        agregar_tabla_resumen_vendedores(ws, row_inicio_mes, ventas_por_vendedor, month_total, mes_nombres[mes], fila_total_mes)
        
        # Espacio entre meses
        current_row += 1
    
    # ⭐ AJUSTAR ANCHOS DE COLUMNAS AL FINAL
    ajustar_anchos_columnas(ws)
    
    # Ajustar altura de filas de encabezados
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20

# ============================================
# HOJA 2: COMISIONES
# ============================================

def crear_hoja_comisiones(wb, ventas, año):
    """Crea la hoja COMISIONES - COPIA EXACTA DE GENERAL con colores naranja/amarillo"""
    ws = wb.create_sheet("Comisiones")
    
    # ⭐ COLORES ESPECÍFICOS PARA COMISIONES
    COLOR_TITULO_COM = "FFA500"        # Naranja
    COLOR_HEADER_COM = "FF8C00"        # Naranja oscuro
    COLOR_MES_COM = "FFD700"           # Amarillo dorado
    COLOR_TOTAL_SEMANAL_COM = "FFE4B5" # Amarillo tenue
    COLOR_TOTAL_MENSUAL_COM = "28A745"  # ⭐ Verde oscuro (mismo que GENERAL)
    
    FONT_TITULO_COM = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    FONT_HEADER_COM = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    FONT_TOTAL_COM = Font(name='Calibri', size=10, bold=True, color='000000')
    FONT_MES_COM = Font(name='Calibri', size=20, bold=True, color='FFFFFF')
    
    # Configurar anchos de columna (IGUAL QUE GENERAL)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 3
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 10
    
    # Fila 2: Título principal
    ws.merge_cells('A2:H2')
    titulo = ws['A2']
    titulo.value = f"Comisiones Anual {año}"
    titulo.font = FONT_TITULO_COM
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    titulo.fill = PatternFill(start_color=COLOR_TITULO_COM, end_color=COLOR_TITULO_COM, fill_type='solid')
    
    # ⭐ Fila 3: Subtítulo
    ws.merge_cells('A3:H3')
    subtitulo = ws['A3']
    subtitulo.value = "Reguladores Electromecánicos, Reparaciones y Mantenimientos"
    subtitulo.font = Font(name='Calibri', size=10, italic=True, color='666666')  # Gris, cursiva
    subtitulo.alignment = Alignment(horizontal='center', vertical='center')
    
    # Filas 5-6: Encabezados en 2 niveles (IGUAL QUE GENERAL)
    # Fecha (A5:A6)
    ws.merge_cells('A5:A6')
    ws.cell(5, 1).value = 'Fecha'
    ws.cell(5, 1).font = FONT_HEADER_COM
    ws.cell(5, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 1).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(5, 1).border = BORDER_THIN
    ws.cell(6, 1).border = BORDER_THIN
    
    # Cliente (B5:B6)
    ws.merge_cells('B5:B6')
    ws.cell(5, 2).value = 'Cliente'
    ws.cell(5, 2).font = FONT_HEADER_COM
    ws.cell(5, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 2).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(5, 2).border = BORDER_THIN
    ws.cell(6, 2).border = BORDER_THIN
    
    # Concepto (C5:C6)
    ws.merge_cells('C5:C6')
    ws.cell(5, 3).value = 'Concepto'
    ws.cell(5, 3).font = FONT_HEADER_COM
    ws.cell(5, 3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 3).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(5, 3).border = BORDER_THIN
    ws.cell(6, 3).border = BORDER_THIN
    
    # DEPÓSITO (D5:E5)
    ws.merge_cells('D5:E5')
    ws.cell(5, 4).value = 'DEPÓSITO'
    ws.cell(5, 4).font = FONT_HEADER_COM
    ws.cell(5, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 4).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(5, 4).border = BORDER_THIN
    ws.cell(5, 5).border = BORDER_THIN
    
    # Total (F5:F6)
    ws.merge_cells('F5:F6')
    ws.cell(5, 6).value = 'Total'
    ws.cell(5, 6).font = FONT_HEADER_COM
    ws.cell(5, 6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 6).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(5, 6).border = BORDER_THIN
    ws.cell(6, 6).border = BORDER_THIN
    
    # Semana (G5:G6)
    ws.merge_cells('G5:G6')
    ws.cell(5, 7).value = 'Semana'
    ws.cell(5, 7).font = FONT_HEADER_COM
    ws.cell(5, 7).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 7).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(5, 7).border = BORDER_THIN
    ws.cell(6, 7).border = BORDER_THIN
    
    # Vendedor (H5:H6)
    ws.merge_cells('H5:H6')
    ws.cell(5, 8).value = 'Vendedor'
    ws.cell(5, 8).font = FONT_HEADER_COM
    ws.cell(5, 8).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(5, 8).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(5, 8).border = BORDER_THIN
    ws.cell(6, 8).border = BORDER_THIN
    
    # Fila 6: Sub-encabezados
    ws.cell(6, 4).value = 'Con IVA'
    ws.cell(6, 4).font = FONT_HEADER_COM
    ws.cell(6, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(6, 4).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(6, 4).border = BORDER_THIN
    
    ws.cell(6, 5).value = 'Sin IVA'
    ws.cell(6, 5).font = FONT_HEADER_COM
    ws.cell(6, 5).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(6, 5).fill = PatternFill(start_color=COLOR_HEADER_COM, end_color=COLOR_HEADER_COM, fill_type='solid')
    ws.cell(6, 5).border = BORDER_THIN
    
    # ⭐ PROCESAMIENTO DE DATOS - IGUAL QUE GENERAL
    current_row = 8
    
    # Agrupar ventas por mes
    ventas_por_mes = {}
    for venta in ventas:
        mes = venta.get('mes', 1)
        if mes not in ventas_por_mes:
            ventas_por_mes[mes] = []
        ventas_por_mes[mes].append(venta)
    
    mes_nombres = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
                   'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    
    # Procesar cada mes
    for mes in range(1, 13):
        if mes not in ventas_por_mes:
            continue
        
        ventas_del_mes = ventas_por_mes[mes]
        
        # ⭐ Título del mes (AMARILLO DORADO con letra NEGRA y BORDES)
        ws.merge_cells(f'A{current_row}:H{current_row}')
        mes_cell = ws.cell(current_row, 1)
        mes_cell.value = mes_nombres[mes]
        mes_cell.font = Font(name='Calibri', size=20, bold=True, color='000000')  # ⭐ NEGRO
        mes_cell.alignment = Alignment(horizontal='center', vertical='center')
        mes_cell.fill = PatternFill(start_color=COLOR_MES_COM, end_color=COLOR_MES_COM, fill_type='solid')
        mes_cell.border = BORDER_THIN  # ⭐ BORDES
        ws.row_dimensions[current_row].height = 30
        
        # ⭐ Agregar bordes a todas las celdas combinadas del título
        for col in range(1, 9):  # A-H
            ws.cell(current_row, col).border = BORDER_THIN
        
        current_row += 1
        
        # Datos de ventas del mes
        current_week = None
        week_total = 0
        month_total = 0
        ventas_por_vendedor = {}
        row_inicio_mes = current_row
        week_start_row = None
        filas_totales_semanales = []  # ⭐ RASTREAR filas de totales semanales para fórmula mensual
        
        for venta in ventas_del_mes:
            # Cambio de semana
            if current_week and venta['week'] != current_week:
                if week_start_row and current_row > week_start_row:
                    # Combinar Semana
                    ws.merge_cells(f'G{week_start_row}:G{current_row - 1}')
                    ws.cell(week_start_row, 7).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # ⭐ Combinar Total con FÓRMULA SUM
                    ws.merge_cells(f'F{week_start_row}:F{current_row - 1}')
                    ws.cell(week_start_row, 6).value = f"=SUM(E{week_start_row}:E{current_row - 1})"  # ⭐ FÓRMULA
                    ws.cell(week_start_row, 6).number_format = '$#,##0.00'
                    ws.cell(week_start_row, 6).font = FONT_TOTAL_COM
                    ws.cell(week_start_row, 6).alignment = Alignment(horizontal='center', vertical='center')
                
                # ⭐ Total semanal (AMARILLO TENUE) con FÓRMULA SUM
                for col in range(1, 9):
                    ws.cell(current_row, col).fill = PatternFill(start_color=COLOR_TOTAL_SEMANAL_COM, end_color=COLOR_TOTAL_SEMANAL_COM, fill_type='solid')
                
                ws.cell(current_row, 4).value = "TOTAL SIN IVA"
                ws.cell(current_row, 4).font = FONT_TOTAL_COM
                ws.cell(current_row, 4).alignment = Alignment(horizontal='right', vertical='center')
                
                ws.cell(current_row, 5).value = f"=SUM(E{week_start_row}:E{current_row - 1})"  # ⭐ FÓRMULA
                ws.cell(current_row, 5).number_format = '$#,##0.00'
                ws.cell(current_row, 5).font = FONT_TOTAL_COM
                
                filas_totales_semanales.append(current_row)  # ⭐ GUARDAR fila de total
                current_row += 1
                week_total = 0
                week_start_row = None
            
            # Nueva semana
            if current_week != venta['week']:
                current_week = venta['week']
                week_start_row = current_row
            
            # Agregar venta
            ws.cell(current_row, 1).value = venta['fecha']
            
            ws.cell(current_row, 2).value = formatear_cliente(venta)
            ws.cell(current_row, 2).alignment = Alignment(wrap_text=True, vertical='top')
            
            concepto_formateado = extraer_info_producto(venta['concepto'])
            ws.cell(current_row, 3).value = concepto_formateado
            ws.cell(current_row, 3).alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.cell(current_row, 4).value = venta['amountWithIVA']
            ws.cell(current_row, 4).number_format = '$#,##0.00'
            
            ws.cell(current_row, 5).value = f"=D{current_row}/1.16"
            ws.cell(current_row, 5).number_format = '$#,##0.00'
            
            if current_row == week_start_row:
                ws.cell(current_row, 7).value = formatear_semana(venta['week'], mes)
            
            vendedor = venta['vendor']
            ws.cell(current_row, 8).value = vendedor
            
            ajustar_altura_fila(ws, current_row)

            # ⭐ Colorear fila si es venta extranjera
            if venta.get('country', 'MX') != 'MX':
                colorear_fila_extranjero(ws, current_row, 8, BORDER_THIN)
            
            amount_sin_iva = venta['amountWithoutIVA']
            week_total += amount_sin_iva
            month_total += amount_sin_iva
            
            if vendedor not in ventas_por_vendedor:
                ventas_por_vendedor[vendedor] = 0
            ventas_por_vendedor[vendedor] += amount_sin_iva
            
            current_row += 1
        
        # Última semana del mes
        if week_start_row and current_row > week_start_row:
            # Combinar Semana
            ws.merge_cells(f'G{week_start_row}:G{current_row - 1}')
            ws.cell(week_start_row, 7).alignment = Alignment(horizontal='center', vertical='center')
            
            # ⭐ Combinar Total con FÓRMULA
            ws.merge_cells(f'F{week_start_row}:F{current_row - 1}')
            ws.cell(week_start_row, 6).value = f"=SUM(E{week_start_row}:E{current_row - 1})"  # ⭐ FÓRMULA
            ws.cell(week_start_row, 6).number_format = '$#,##0.00'
            ws.cell(week_start_row, 6).font = FONT_TOTAL_COM
            ws.cell(week_start_row, 6).alignment = Alignment(horizontal='center', vertical='center')
        
        if week_total > 0:
            # ⭐ Total semanal con FÓRMULA
            for col in range(1, 9):
                ws.cell(current_row, col).fill = PatternFill(start_color=COLOR_TOTAL_SEMANAL_COM, end_color=COLOR_TOTAL_SEMANAL_COM, fill_type='solid')
            
            ws.cell(current_row, 4).value = "TOTAL SIN IVA"
            ws.cell(current_row, 4).font = FONT_TOTAL_COM
            ws.cell(current_row, 4).alignment = Alignment(horizontal='right', vertical='center')
            
            ws.cell(current_row, 5).value = f"=SUM(E{week_start_row}:E{current_row - 1})"  # ⭐ FÓRMULA
            ws.cell(current_row, 5).number_format = '$#,##0.00'
            ws.cell(current_row, 5).font = FONT_TOTAL_COM
            
            filas_totales_semanales.append(current_row)  # ⭐ GUARDAR fila de total
            current_row += 1
        
        # ⭐ Total del mes (VERDE) con FÓRMULA
        ws.cell(current_row, 4).value = f"VENTAS {mes_nombres[mes]}"
        ws.cell(current_row, 4).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws.cell(current_row, 4).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(current_row, 4).fill = PatternFill(start_color=COLOR_TOTAL_MENSUAL_COM, end_color=COLOR_TOTAL_MENSUAL_COM, fill_type='solid')
        
        # ⭐ FÓRMULA: Sumar todas las filas de totales semanales
        if filas_totales_semanales:
            # Construir fórmula sumando celdas E de cada total semanal
            celdas_sum = "+".join([f"E{fila}" for fila in filas_totales_semanales])
            ws.cell(current_row, 5).value = f"={celdas_sum}"
        else:
            ws.cell(current_row, 5).value = 0
        
        ws.cell(current_row, 5).number_format = '$#,##0.00'
        ws.cell(current_row, 5).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws.cell(current_row, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(current_row, 5).fill = PatternFill(start_color=COLOR_TOTAL_MENSUAL_COM, end_color=COLOR_TOTAL_MENSUAL_COM, fill_type='solid')
        
        fila_total_mes = current_row  # ⭐ Guardar fila para tabla
        current_row += 1
        
        # ⭐ Tabla de COMISIONES (nueva estructura)
        agregar_tabla_comisiones_mes(ws, row_inicio_mes, fila_total_mes, mes_nombres[mes])
        
        # Espacio entre meses
        current_row += 1
    
    ajustar_anchos_columnas(ws)

# ============================================
# FUNCIONES AUXILIARES
# ============================================

def agregar_total_semanal_general(ws, row, week_start_row, filas_totales_semanales):
    """Agrega fila de total semanal en GENERAL con FÓRMULA"""
    # ⭐ FONDO ROJO TENUE EN TODA LA FILA (A-H) para delimitador visual
    for col in range(1, 9):  # Columnas A-H
        ws.cell(row, col).fill = PatternFill(start_color='FFB3BA', end_color='FFB3BA', fill_type='solid')  # ⭐ Rosa/Rojo tenue
    
    # ⭐ TEXTO "TOTAL SIN IVA" en columna D
    ws.cell(row, 4).value = "TOTAL SIN IVA"
    ws.cell(row, 4).font = Font(name='Calibri', size=10, bold=True, color='000000')  # ⭐ Negro para contrastar
    ws.cell(row, 4).alignment = Alignment(horizontal='right', vertical='center')
    
    # ⭐ FÓRMULA en columna E: SUM desde inicio de semana hasta fila anterior
    ws.cell(row, 5).value = f"=SUM(E{week_start_row}:E{row - 1})"  # ⭐ FÓRMULA
    ws.cell(row, 5).number_format = '$#,##0.00'
    ws.cell(row, 5).font = Font(name='Calibri', size=10, bold=True, color='000000')  # ⭐ Negro
    
    filas_totales_semanales.append(row)  # ⭐ GUARDAR fila para fórmula mensual
    
    return row + 1


def agregar_total_mensual_general(ws, row, total_mes, mes_nombre):
    """Agrega fila de total mensual en GENERAL"""
    # ⭐ TEXTO "VENTAS ENERO" en columna D (al lado de la cifra)
    ws.cell(row, 4).value = f"VENTAS {mes_nombre}"
    ws.cell(row, 4).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # ⭐ Tamaño 11
    ws.cell(row, 4).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row, 4).fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')  # ⭐ Verde
    
    # Total del mes en columna E
    ws.cell(row, 5).value = total_mes
    ws.cell(row, 5).number_format = '$#,##0.00'
    ws.cell(row, 5).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # ⭐ Tamaño 11
    ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row, 5).fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')  # ⭐ Verde
    
    return row + 1


def agregar_tabla_resumen_vendedores(ws, row_inicio, ventas_por_vendedor, total_mes, mes_nombre, fila_total_mes):
    """
    Agrega tabla resumen de ventas por vendedor para GENERAL (columnas J-K-L)
    """
    col_vendedor = 10  # J
    col_ventas = 11    # K
    col_porcentaje = 12  # L
    
    current_row = row_inicio
    
    # Título de la tabla
    ws.merge_cells(f'J{current_row}:L{current_row}')
    titulo = ws.cell(current_row, col_vendedor)
    titulo.value = mes_nombre
    titulo.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    titulo.fill = PatternFill(start_color='F62E41', end_color='F62E41', fill_type='solid')  # Rojo REVOLT
    current_row += 1
    
    # Encabezados de la tabla
    headers = ['VENDEDOR', 'VENTAS', '%']
    for idx, col in enumerate([col_vendedor, col_ventas, col_porcentaje]):
        cell = ws.cell(current_row, col)
        cell.value = headers[idx]
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D41F33', end_color='D41F33', fill_type='solid')  # Rojo oscuro
        cell.border = BORDER_THIN
    current_row += 1
    
    # Datos por vendedor
    for vendedor, monto in sorted(ventas_por_vendedor.items()):
        # Vendedor
        ws.cell(current_row, col_vendedor).value = vendedor
        ws.cell(current_row, col_vendedor).font = FONT_NORMAL
        ws.cell(current_row, col_vendedor).border = BORDER_THIN
        
        # ⭐ Ventas con FÓRMULA SUMIF
        ws.cell(current_row, col_ventas).value = f'=SUMIF(H{row_inicio}:H{fila_total_mes-1},J{current_row},E{row_inicio}:E{fila_total_mes-1})'
        ws.cell(current_row, col_ventas).number_format = '$#,##0.00'
        ws.cell(current_row, col_ventas).font = FONT_NORMAL
        ws.cell(current_row, col_ventas).border = BORDER_THIN
        
        # ⭐ Porcentaje con FÓRMULA
        ws.cell(current_row, col_porcentaje).value = f"=K{current_row}/E${fila_total_mes}"
        ws.cell(current_row, col_porcentaje).number_format = '0.00%'
        ws.cell(current_row, col_porcentaje).font = FONT_NORMAL
        ws.cell(current_row, col_porcentaje).border = BORDER_THIN
        
        current_row += 1
    
    # Fila de TOTAL
    ws.cell(current_row, col_vendedor).value = "Total"
    ws.cell(current_row, col_vendedor).font = FONT_TOTAL
    ws.cell(current_row, col_vendedor).border = BORDER_THIN
    ws.cell(current_row, col_vendedor).fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    
    ws.cell(current_row, col_ventas).value = f"=E{fila_total_mes}"
    ws.cell(current_row, col_ventas).number_format = '$#,##0.00'
    ws.cell(current_row, col_ventas).font = FONT_TOTAL
    ws.cell(current_row, col_ventas).border = BORDER_THIN
    ws.cell(current_row, col_ventas).fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    
    ws.cell(current_row, col_porcentaje).value = 1.0  # 100%
    ws.cell(current_row, col_porcentaje).number_format = '0.00%'
    ws.cell(current_row, col_porcentaje).font = FONT_TOTAL
    ws.cell(current_row, col_porcentaje).border = BORDER_THIN
    ws.cell(current_row, col_porcentaje).fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
    
    return current_row + 1


def agregar_tabla_comisiones_mes(ws, row_inicio, fila_total_mes, mes_nombre):
    """
    Agrega tabla de comisiones por mes con estructura específica
    
    Estructura:
    ┌────────────────────────────────┐
    │  VENTAS ENERO                  │ ← Título
    ├────────────────────────────────┤
    │         $1,500,000.00          │ ← Cifra grande (2 filas combinadas)
    │                                │
    ├──────────────┬─────────────────┤
    │  Vendedor    │   Comisiones    │ ← Headers
    ├──────────────┼─────────────────┤
    │  Hugo        │   $33,300.00    │ ← 2.22% del total
    │  Auxiliar    │   $22,200.00    │ ← 1.48% del total
    └──────────────┴─────────────────┘
    
    Rangos de comisión:
    - $0 - $690,000: 0% (Hugo: 0%, Auxiliar: 0%)
    - $690,001 - $890,000: 1.50% (Hugo: 1%, Auxiliar: 0.5%)
    - $890,001 - $1,090,000: 2.50% (Hugo: 1.5%, Auxiliar: 1%)
    - $1,090,001 - $1,290,000: 3.50% (Hugo: 2.1%, Auxiliar: 1.4%)
    - $1,290,001 - $2,500,000: 3.70% (Hugo: 2.22%, Auxiliar: 1.48%)
    """
    
    col_vendedor = 10  # J
    col_comisiones = 11  # K
    
    current_row = row_inicio
    
    # ⭐ FILA 1: Título "VENTAS ENERO"
    ws.merge_cells(f'J{current_row}:K{current_row}')
    titulo = ws.cell(current_row, col_vendedor)
    titulo.value = f"VENTAS {mes_nombre}"
    titulo.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    titulo.fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')  # Naranja oscuro
    titulo.border = BORDER_THIN
    ws.cell(current_row, col_comisiones).border = BORDER_THIN
    current_row += 1
    
    # ⭐ FILA 2-3: Cifra total (COMBINADA VERTICALMENTE, GRANDE)
    ws.merge_cells(f'J{current_row}:K{current_row + 1}')
    cifra = ws.cell(current_row, col_vendedor)
    cifra.value = f"=E{fila_total_mes}"  # ⭐ FÓRMULA referencia al total mensual
    cifra.number_format = '$#,##0.00'
    cifra.font = Font(name='Calibri', size=16, bold=True, color='000000')  # ⭐ Grande
    cifra.alignment = Alignment(horizontal='center', vertical='center')
    cifra.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')  # Amarillo tenue
    cifra.border = BORDER_THIN
    ws.cell(current_row, col_comisiones).border = BORDER_THIN
    ws.cell(current_row + 1, col_vendedor).border = BORDER_THIN
    ws.cell(current_row + 1, col_comisiones).border = BORDER_THIN
    current_row += 2
    
    # ⭐ FILA 4: Headers (Vendedor | Comisiones)
    headers = ['Vendedor', 'Comisiones']
    for idx, col in enumerate([col_vendedor, col_comisiones]):
        cell = ws.cell(current_row, col)
        cell.value = headers[idx]
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Naranja
        cell.border = BORDER_THIN
    current_row += 1
    
    # ⭐ FILAS 5-6: Hugo y Marlen con FÓRMULAS de comisión
    vendedores = [
        ('Hugo', 'hugo'),
        ('Marlen', 'marlen')  # ⭐ Cambio: Auxiliar → Marlen
    ]
    
    for nombre, tipo in vendedores:
        # Vendedor
        ws.cell(current_row, col_vendedor).value = nombre
        ws.cell(current_row, col_vendedor).font = Font(name='Calibri', size=10, bold=False)
        ws.cell(current_row, col_vendedor).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(current_row, col_vendedor).border = BORDER_THIN
        
        # ⭐ Comisión con FÓRMULA IF anidados (en lugar de IFS para compatibilidad)
        # Celda del total: J{row_inicio + 1} (donde está la cifra combinada)
        celda_total = f"J{row_inicio + 1}"
        
        if tipo == 'hugo':
            # Hugo: 0%, 1%, 1.5%, 2.1%, 2.22%
            formula = f'=IF({celda_total}<=690000,0,IF({celda_total}<=890000,{celda_total}*0.01,IF({celda_total}<=1090000,{celda_total}*0.015,IF({celda_total}<=1290000,{celda_total}*0.021,{celda_total}*0.0222))))'
        else:  # marlen
            # Marlen: 0%, 0.5%, 1%, 1.4%, 1.48%
            formula = f'=IF({celda_total}<=690000,0,IF({celda_total}<=890000,{celda_total}*0.005,IF({celda_total}<=1090000,{celda_total}*0.01,IF({celda_total}<=1290000,{celda_total}*0.014,{celda_total}*0.0148))))'
        
        ws.cell(current_row, col_comisiones).value = formula
        ws.cell(current_row, col_comisiones).number_format = '$#,##0.00'
        ws.cell(current_row, col_comisiones).font = Font(name='Calibri', size=10, bold=False)
        ws.cell(current_row, col_comisiones).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(current_row, col_comisiones).border = BORDER_THIN
        
        current_row += 1
    
    return current_row + 1
def agregar_total_mensual_comisiones(ws, row, data_mes, mes_nombre):
    """Agrega fila de total mensual en COMISIONES"""
    # ⭐ TEXTO "VENTAS ENERO" en columna D
    ws.cell(row, 4).value = f"VENTAS {mes_nombre}"
    ws.cell(row, 4).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # ⭐ Tamaño 11
    ws.cell(row, 4).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row, 4).fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')  # ⭐ Verde
    
    # Total sin IVA del mes
    ws.cell(row, 5).value = data_mes['total_sin_iva']
    ws.cell(row, 5).number_format = '$#,##0.00'
    ws.cell(row, 5).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # ⭐ Tamaño 11
    ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row, 5).fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    
    # Comisión Hugo del mes
    ws.cell(row, 13).value = data_mes['comision_hugo']
    ws.cell(row, 13).number_format = '$#,##0.00'
    ws.cell(row, 13).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # ⭐ Tamaño 11
    ws.cell(row, 13).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row, 13).fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    
    # Comisión Auxiliar del mes
    ws.cell(row, 14).value = data_mes['comision_aux']
    ws.cell(row, 14).number_format = '$#,##0.00'
    ws.cell(row, 14).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # ⭐ Tamaño 11
    ws.cell(row, 14).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row, 14).fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    
    return row + 1

# ============================================
# MAIN
# ============================================

def crear_hoja_comercializacion(wb, ventas, año):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Comercializacion")

    COLOR_TITULO_PRINCIPAL = "0070C0"
    COLOR_HEADER_FILA      = "1F4E79"
    COLOR_HUGO             = "DDEBF7"
    COLOR_MARLEN           = "FCE4D6"
    COLOR_TOTAL_HUGO       = "BDD7EE"
    COLOR_TOTAL_MARLEN     = "F4B8A0"

    FONT_TITULO_DOC = Font(name='Calibri', size=16, bold=True)
    FONT_SUBTITULO  = Font(name='Calibri', size=12, color='222222')
    FONT_MES        = Font(name='Calibri', size=16, bold=True, color='FFFFFF', italic=True)
    FONT_HEADER     = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    FONT_NORMAL     = Font(name='Calibri', size=10)
    FONT_TOTAL      = Font(name='Calibri', size=10, bold=True)
    FONT_VENDEDOR   = Font(name='Calibri', size=10, bold=True)

    THIN   = Side(style='thin', color='000000')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 13

    ws.merge_cells('A1:J1')
    ws['A1'].value = f"Comisions Mensual Equipo de Comercializacion {año}"
    ws['A1'].font = FONT_TITULO_DOC
    ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[1].height = 20

    ws.merge_cells('A2:J2')
    ws['A2'].value = "Reguladores Electrónicos, Equipos de EC, UPS, No Breaks, Plantas, Transformadores, Instalaciones"
    ws['A2'].font = FONT_SUBTITULO
    ws['A2'].alignment = Alignment(horizontal='right', vertical='center')

    ventas_por_mes = {}
    for venta in ventas:
        mes = venta.get('mes', 1)
        if mes not in ventas_por_mes:
            ventas_por_mes[mes] = []
        ventas_por_mes[mes].append(venta)

    mes_nombres = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
                   'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

    current_row = 4

    for mes in range(1, 13):
        if mes not in ventas_por_mes:
            continue

        ventas_del_mes = ventas_por_mes[mes]
        ventas_hugo   = [v for v in ventas_del_mes if 'hugo' in v.get('vendor', '').lower()]
        ventas_marlen = [v for v in ventas_del_mes if 'hugo' not in v.get('vendor', '').lower()]
        if not ventas_hugo and not ventas_marlen:
            ventas_hugo = ventas_del_mes

        # Título mes
        ws.merge_cells(f'A{current_row}:J{current_row}')
        c = ws.cell(current_row, 1)
        c.value = f"Comisiones Equipos de Comercializacion {mes_nombres[mes]}"
        c.font = FONT_MES
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.fill = PatternFill(start_color=COLOR_TITULO_PRINCIPAL, end_color=COLOR_TITULO_PRINCIPAL, fill_type='solid')
        ws.row_dimensions[current_row].height = 21
        for col in range(1, 11):
            ws.cell(current_row, col).border = BORDER
        current_row += 1

        # Headers nivel 1
        for col, label in [(1,'Fecha'), (2,'Cliente'), (3,'Concepto')]:
            ws.merge_cells(f'{get_column_letter(col)}{current_row}:{get_column_letter(col)}{current_row+1}')
            c = ws.cell(current_row, col)
            c.value = label
            c.font = FONT_HEADER
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color=COLOR_HEADER_FILA, end_color=COLOR_HEADER_FILA, fill_type='solid')
            c.border = BORDER
            ws.cell(current_row+1, col).border = BORDER

        for col_start, col_end, label in [(4,5,'Deposito'), (6,7,'Precio Proveedor')]:
            ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_end)}{current_row}')
            c = ws.cell(current_row, col_start)
            c.value = label
            c.font = FONT_HEADER
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color=COLOR_HEADER_FILA, end_color=COLOR_HEADER_FILA, fill_type='solid')
            c.border = BORDER
            ws.cell(current_row, col_start+1).border = BORDER

        for col, label in [(8,'Base para\nComisión'), (9,'Comision'), (10,'Vendedor')]:
            ws.merge_cells(f'{get_column_letter(col)}{current_row}:{get_column_letter(col)}{current_row+1}')
            c = ws.cell(current_row, col)
            c.value = label
            c.font = FONT_HEADER
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.fill = PatternFill(start_color=COLOR_HEADER_FILA, end_color=COLOR_HEADER_FILA, fill_type='solid')
            c.border = BORDER
            ws.cell(current_row+1, col).border = BORDER
        current_row += 1

        # Headers nivel 2
        for col, label in [(4,'Con IVA'), (5,'Sin IVA'), (6,'Con IVA'), (7,'Sin IVA')]:
            c = ws.cell(current_row, col)
            c.value = label
            c.font = FONT_HEADER
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color=COLOR_HEADER_FILA, end_color=COLOR_HEADER_FILA, fill_type='solid')
            c.border = BORDER
        current_row += 1

        # Bloque Hugo
        hugo_start = current_row
        for venta in ventas_hugo:
            _agregar_fila_comercializacion(ws, current_row, venta, FONT_NORMAL, BORDER, COLOR_HUGO, 0.05)
            current_row += 1
        if not ventas_hugo:
            _agregar_fila_comercializacion(ws, current_row, None, FONT_NORMAL, BORDER, COLOR_HUGO, 0.05)
            current_row += 1

        hugo_total_row = current_row
        _agregar_total_comercializacion(ws, current_row, hugo_start, current_row - 1, COLOR_TOTAL_HUGO, FONT_TOTAL, BORDER)
        current_row += 1

        if hugo_total_row > hugo_start:
            ws.merge_cells(f'J{hugo_start}:J{hugo_total_row}')
        c = ws.cell(hugo_start, 10)
        c.value = 'Hugo'
        c.font = FONT_VENDEDOR
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill(start_color=COLOR_HUGO, end_color=COLOR_HUGO, fill_type='solid')
        c.border = BORDER

        # Bloque Marlen
        marlen_start = current_row
        for venta in ventas_marlen:
            _agregar_fila_comercializacion(ws, current_row, venta, FONT_NORMAL, BORDER, COLOR_MARLEN, 0.10)
            current_row += 1
        if not ventas_marlen:
            _agregar_fila_comercializacion(ws, current_row, None, FONT_NORMAL, BORDER, COLOR_MARLEN, 0.10)
            current_row += 1

        marlen_total_row = current_row
        _agregar_total_comercializacion(ws, current_row, marlen_start, current_row - 1, COLOR_TOTAL_MARLEN, FONT_TOTAL, BORDER)
        current_row += 1

        if marlen_total_row > marlen_start:
            ws.merge_cells(f'J{marlen_start}:J{marlen_total_row}')
        c = ws.cell(marlen_start, 10)
        c.value = 'Marlen'
        c.font = FONT_VENDEDOR
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill(start_color=COLOR_MARLEN, end_color=COLOR_MARLEN, fill_type='solid')
        c.border = BORDER

        current_row += 1


def _agregar_fila_comercializacion(ws, row, venta, font, border, color_fill, pct_comision):
    from openpyxl.styles import Alignment, PatternFill
    fill = PatternFill(start_color=color_fill, end_color=color_fill, fill_type='solid')

    if venta is None:
        for col in range(1, 11):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).border = border
        return

    c = ws.cell(row, 1)
    c.value = venta.get('fecha', '')
    c.font = font; c.fill = fill; c.border = border
    c.alignment = Alignment(horizontal='center', vertical='center')

    cliente = venta.get('cliente', '')
    empresa = venta.get('empresa', '')
    c = ws.cell(row, 2)
    c.value = f"{cliente} - {empresa}" if empresa else cliente
    c.font = font; c.fill = fill; c.border = border
    c.alignment = Alignment(wrap_text=True, vertical='top')

    c = ws.cell(row, 3)
    c.value = venta.get('concepto', '')
    c.font = font; c.fill = fill; c.border = border
    c.alignment = Alignment(wrap_text=True, vertical='top')

    c = ws.cell(row, 4)
    c.value = venta.get('amountWithIVA', 0)
    c.number_format = '$#,##0.00'; c.font = font; c.fill = fill; c.border = border

    c = ws.cell(row, 5)
    c.value = f"=D{row}/1.16"
    c.number_format = '$#,##0.00'; c.font = font; c.fill = fill; c.border = border

    c = ws.cell(row, 6)
    c.value = venta.get('providerCostWithIVA', 0)
    c.number_format = '$#,##0.00'; c.font = font; c.fill = fill; c.border = border

    c = ws.cell(row, 7)
    c.value = f"=F{row}/1.16"
    c.number_format = '$#,##0.00'; c.font = font; c.fill = fill; c.border = border

    c = ws.cell(row, 8)
    c.value = f"=E{row}-G{row}"
    c.number_format = '$#,##0.00'; c.font = font; c.fill = fill; c.border = border

    c = ws.cell(row, 9)
    c.value = f"=H{row}*{pct_comision}"
    c.number_format = '$#,##0.00'; c.font = font; c.fill = fill; c.border = border
    # ⭐ Colorear fila si es venta extranjera
    if venta.get('country', 'MX') != 'MX':
        colorear_fila_extranjero(ws, row, 9, border)


def _agregar_total_comercializacion(ws, row, start_row, end_row, color_total, font_total, border):
    from openpyxl.styles import Alignment, PatternFill
    fill = PatternFill(start_color=color_total, end_color=color_total, fill_type='solid')

    for col in range(1, 11):
        ws.cell(row, col).fill = fill
        ws.cell(row, col).border = border

    c = ws.cell(row, 8)
    c.value = 'TOTAL'
    c.font = font_total
    c.alignment = Alignment(horizontal='right', vertical='center')

    c = ws.cell(row, 9)
    c.value = f"=SUM(I{start_row}:I{end_row})"
    c.number_format = '$#,##0.00'
    c.font = font_total

def colorear_fila_extranjero(ws, row, num_cols, border):
    """Si la venta es extranjera, pinta la fila de azul claro"""
    from openpyxl.styles import PatternFill
    COLOR_EXTRANJERO = "DDEEFF"
    fill = PatternFill(start_color=COLOR_EXTRANJERO, end_color=COLOR_EXTRANJERO, fill_type='solid')
    for col in range(1, num_cols + 1):
        cell = ws.cell(row, col)
        # Solo pintar si no tiene ya un color especial (totales, encabezados)
        if cell.fill.fill_type == 'none' or cell.fill.fgColor.rgb in ['00000000', 'FFFFFFFF', COLOR_EXTRANJERO]:
            cell.fill = fill
            cell.border = border

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("[ERROR] Uso: python script.py <json_file> <output_path> <año>")
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_path = sys.argv[2]
    año = int(sys.argv[3])
    
    # Leer JSON
    with open(json_file, 'r', encoding='utf-8') as f:
        data_json = f.read()
    
    generar_reporte(data_json, output_path, año)