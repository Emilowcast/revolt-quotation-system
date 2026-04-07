#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
generate-production-order.py
Genera Orden de Produccion (Regulador) llenando el template Excel REV 7
Uso: python generate-production-order.py <json_file> <template_path> <output_path>
"""

import sys
import json
import re
import os
import shutil
import math
from datetime import datetime

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

def extraer_capacidad(modelo, descripcion):
    texto = f"{modelo} {descripcion}".upper()
    m = re.search(r'(\d+\.?\d*)\s*KVA', texto)
    if m:
        val = float(m.group(1))
        return int(val) if val == int(val) else val
    m = re.search(r'RM-(\d{2,3})', modelo.upper())
    if m:
        num = m.group(1)
        try:
            return int(num[:2] if len(num) == 3 else num[:1])
        except:
            pass
    return None

def extraer_fases(modelo, descripcion):
    texto = f"{modelo} {descripcion}".lower()
    # Detectar primero los casos más específicos
    if re.search(r'bif[aá]sico\s*con\s*neutro|2fn|2\s*f\s*n', texto):
        return '2FN'
    if re.search(r'trif[aá]sico|trifas|3\s*f\b|3f\b', texto):
        return '3F'
    if re.search(r'bif[aá]sico\s*sin\s*neutro|bif[aá]sico|bifas|2\s*f\b|2f\b', texto):
        return '2F'
    if re.search(r'monof[aá]sico|monofas|1\s*f\b|1f\b', texto):
        return '1F'
    return '1F'

def fases_a_numero(tipo_fase):
    # 2F sin neutro se considera monofasico (1) para calculo de amperaje
    return {'1F': 1, '2F': 1, '2FN': 2, '3F': 3}.get(tipo_fase, 1)

def extraer_voltaje_salida(modelo, descripcion):
    m = re.search(r'-(\d{3})(?:\s|$)', modelo.upper())
    if m:
        return int(m.group(1))
    m = re.search(r'(\d{3})\s*v(?:oltios?)?', descripcion.lower())
    if m:
        return int(m.group(1))
    return None

def split_texto(texto, n=3):
    if not texto:
        return [None] * n
    # Soportar separación por salto de línea O por coma/punto y coma
    partes = [p.strip() for p in re.split(r'\n|[,;]', texto) if p.strip()]
    while len(partes) < n:
        partes.append(None)
    return partes[:n]

def calcular_amperaje(capacidad, num_fases, voltaje):
    try:
        kva = float(capacidad)
        f   = float(num_fases) if num_fases else 1
        v   = float(voltaje)
        if f == 1:
            return round((kva * 1000) / v, 2)
        else:
            return round((kva * 1000 / f) / (v / math.sqrt(3)), 2)
    except:
        return None

def generar_op(data, template_path, output_path):
    import openpyxl

    modelo        = data.get('modelo', '')
    descripcion   = data.get('descripcion', '')
    folio         = data.get('folio', '')
    cliente       = data.get('clientName', '')
    empresa       = data.get('clientCompany', '')
    cant          = data.get('cant', 1)
    es_transformador = data.get('esTransformador', False)

    # Cliente completo
    if empresa and cliente:
        cliente_display = f"{cliente} - {empresa}"
    elif empresa:
        cliente_display = empresa
    else:
        cliente_display = cliente

    capacidad  = extraer_capacidad(modelo, descripcion)
    tipo_fase  = extraer_fases(modelo, descripcion)
    num_fases  = fases_a_numero(tipo_fase)

    def to_num(val):
        try:
            return float(val) if val not in ('', None) else None
        except:
            return None

    if es_transformador:
        volt_entrada = to_num(data.get('voltajeEntrada'))
        volt_sal     = to_num(data.get('voltajeSalida')) or to_num(extraer_voltaje_salida(modelo, descripcion))
        volt_min     = None
        volt_max     = None
    else:
        volt_min     = to_num(data.get('voltajeMinEntrada'))
        volt_max     = to_num(data.get('voltajeMaxEntrada'))
        volt_sal     = to_num(data.get('voltajeSalida')) or to_num(extraer_voltaje_salida(modelo, descripcion))
        volt_entrada = None

    if es_transformador:
        amp_entrada = calcular_amperaje(capacidad, num_fases, volt_entrada) if volt_entrada and capacidad else None
        amp_sal     = calcular_amperaje(capacidad, num_fases, volt_sal) if volt_sal and capacidad else None
        amp_min     = None
        amp_max     = None
    else:
        amp_min     = calcular_amperaje(capacidad, num_fases, volt_min) if volt_min and capacidad else None
        amp_max     = calcular_amperaje(capacidad, num_fases, volt_max) if volt_max and capacidad else None
        amp_sal     = calcular_amperaje(capacidad, num_fases, volt_sal) if volt_sal and capacidad else None
        amp_entrada = None

    meses = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']
    hoy = datetime.now()
    fecha_hoy = f"{hoy.day:02d}-{meses[hoy.month-1]}-{str(hoy.year)[2:]}"

    adicionales   = split_texto(data.get('adicionales', ''), 9)
    observaciones = split_texto(data.get('observaciones', ''), 3)

    marca = {
        '1F':  ('X', '.', '.', '.'),
        '2F':  ('.', 'X', '.', '.'),
        '2FN': ('.', '.', 'X', '.'),
        '3F':  ('.', '.', '.', 'X')
    }
    m1f, m2f, m2fn, m3f = marca.get(tipo_fase, ('X', '.', '.', '.'))

    numeros_texto = ['', 'UNO', 'DOS', 'TRES', 'CUATRO', 'CINCO']
    try:
        cant_int   = int(cant)
        cant_label = f"{cant_int} ({numeros_texto[cant_int]})" if cant_int < len(numeros_texto) else str(cant_int)
    except:
        cant_label = str(cant)

    # Copiar template
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    # ⭐ Usar template según tipo de producto
    if es_transformador:
        template_dir  = os.path.dirname(template_path)
        template_real = os.path.join(template_dir, 'FORMATO_OP_TRANSFORMADOR.xlsx')
        if not os.path.exists(template_real):
            template_real = template_path  # fallback al regulador si no existe
    else:
        template_real = template_path
    shutil.copy2(template_real, output_path)

    wb = openpyxl.load_workbook(output_path, data_only=False, keep_links=False)
    nombre_hoja = 'ORDEN DE PRODUCCION TRANSF' if es_transformador else 'ORDEN DE PRODUCCION REV 7'
    ws = wb[nombre_hoja]

    def sv(coord, value):
        if value is None:
            return
        ws[coord].value = value

    def clear(coord):
        """Limpiar celda (eliminar datos de ejemplo del template)."""
        ws[coord].value = None

    # ══ Limpiar TODOS los datos de ejemplo del template ══
    clear('W13')   # folio de ejemplo
    clear('X13')   # descripcion de ejemplo
    clear('U16')   # numero de serie de ejemplo
    clear('U17')   # fecha salida de ejemplo
    clear('U18')   # adicional de ejemplo (TOMAR STOCK PARA ENVIO)
    clear('U19')   # adicional de ejemplo (CON JAULA DE EMBARQUE)
    clear('U20')   # adicional de ejemplo
    # Numero de serie: K10 y K37 referencian U16
    ws['K10'].value = '=IF(U16="","",U16)'
    ws['K37'].value = '=IF(U16="","",U16)'
    # Fecha salida: G14 y G41 referencian U17 con formato fecha
    ws['G14'].value = '=IF(U17="","",U17)'
    ws['G41'].value = '=IF(U17="","",U17)'

    # ══ ZONA DE DATOS — Columna U ══
    # (numero de serie U16 y fecha salida U17 se dejan en blanco — ingreso manual en Excel)
    # ⭐ En la OP solo se plasma el número del folio, no el prefijo completo
    numero_folio = folio.split('-')[-1] if folio else folio
    sv('U1',  numero_folio)
    sv('U2',  cliente_display)
    sv('U3',  'GABINETE TRANSFORMADOR' if es_transformador else 'REGULADOR')
    sv('U4',  capacidad)
    sv('U6',  num_fases)
    if es_transformador:
        sv('U7',  cant_label)    # En transformador U7 = cantidad
        sv('U8',  volt_entrada)  # Voltaje entrada directo
        sv('U9',  volt_sal)      # Voltaje salida directo
    else:
        sv('U7',  volt_min)
        sv('U8',  volt_max)
        sv('U9',  volt_sal)
    sv('U10', modelo)
    sv('U11', m1f)
    sv('U13', m2f)
    sv('U14', m2fn)
    sv('U15', m3f)
    # Adicionales → U18 a U26 (saltando celdas fusionadas U21 y U25)
    celdas_adicionales_u = ['U18','U19','U20','U22','U23','U24','U26']
    for i, celda in enumerate(celdas_adicionales_u):
        cell = ws[celda]
        if cell.__class__.__name__ != 'MergedCell':
            sv(celda, adicionales[i] if i < len(adicionales) else None)

    # Observaciones → U28 a U30 (3 líneas, verificando fusión)
    for i, celda in enumerate(['U28','U29','U30']):
        cell = ws[celda]
        if cell.__class__.__name__ != 'MergedCell':
            sv(celda, observaciones[i] if i < len(observaciones) else None)

    # ══ TABLA 1 y TABLA 2 — Solo escribir valores directos que NO tienen formula =U* ══
    # (el resto se propaga automáticamente desde columna U via formulas del template)
    ws['P3'].value = '=U1'
    sv('N3',  fecha_hoy)
    sv('M5',  cliente_display)
    sv('A8',  'GABINETE TRANSFORMADOR' if es_transformador else 'REGULADOR')
    sv('C8',  capacidad)
    if es_transformador:
        sv('E8',  volt_entrada)
        sv('H8',  volt_sal)
        sv('I8',  amp_entrada)
        sv('L8',  amp_sal)
    else:
        sv('E8',  volt_min)
        sv('G8',  volt_max)
        sv('H8',  volt_sal)
        sv('I8',  amp_min)
        sv('K8',  amp_max)
        sv('L8',  amp_sal)
    sv('M8',  modelo)
    sv('N8',  cant_label)
    sv('B10', m1f)
    sv('D10', m2f)
    sv('G10', m2fn)
    sv('I10', m3f)
    sv('G12', fecha_hoy)
    # TABLA 2
    sv('N30', fecha_hoy)
    ws['P30'].value = '=U1'
    sv('M32', cliente_display)
    sv('A35', 'GABINETE TRANSFORMADOR' if es_transformador else 'REGULADOR')
    sv('C35', capacidad)
    if es_transformador:
        sv('E35', volt_entrada)
        sv('H35', volt_sal)
        sv('I35', amp_entrada)
        sv('L35', amp_sal)
    else:
        sv('E35', volt_min)
        sv('G35', volt_max)
        sv('H35', volt_sal)
        sv('I35', amp_min)
        sv('K35', amp_max)
        sv('L35', amp_sal)
    sv('M35', modelo)
    sv('N35', cant_label)
    sv('B37', m1f)
    sv('D37', m2f)
    sv('G37', m2fn)
    sv('I37', m3f)
    sv('G39', fecha_hoy)

    # ══ Tabla 1 — Adicionales K13 a K19 con fórmulas de igualación ══
    celdas_adic_t1 = [('K13','U18'),('K14','U19'),('K15','U20'),('K16','U22'),
                      ('K17','U23'),('K18','U24'),('K19','U26')]
    for celda_k, celda_u in celdas_adic_t1:
        cell = ws[celda_k]
        if cell.__class__.__name__ != 'MergedCell':
            cell.value = f'=IF({celda_u}="","",{celda_u})'

    # ══ Tabla 1 — Observaciones A17 a A19 con fórmulas de igualación ══
    obs_map_t1 = [('A17','U28'),('A18','U29'),('A19','U30')]
    for celda_a, celda_u in obs_map_t1:
        cell = ws[celda_a]
        if cell.__class__.__name__ != 'MergedCell':
            cell.value = f'=IF({celda_u}="","",{celda_u})'

    # ══ Tabla 2 — Adicionales K40 a K46 con fórmulas de igualación ══
    celdas_adic_t2 = [('K40','K13'),('K41','K14'),('K42','K15'),('K43','K16'),
                      ('K44','K17'),('K45','K18'),('K46','K19')]
    for celda_k2, celda_k1 in celdas_adic_t2:
        cell = ws[celda_k2]
        if cell.__class__.__name__ != 'MergedCell':
            cell.value = f'=IF({celda_k1}="","",{celda_k1})'

    # ══ Tabla 2 — Observaciones A44 a A46 con fórmulas de igualación ══
    obs_map_t2 = [('A44','A17'),('A45','A18'),('A46','A19')]
    for celda_a2, celda_a1 in obs_map_t2:
        cell = ws[celda_a2]
        if cell.__class__.__name__ != 'MergedCell':
            cell.value = f'=IF({celda_a1}="","",{celda_a1})'

    wb.save(output_path)
    print(f"OK: Orden generada: {output_path}")
    print(f"OK: Orden generada: {output_path}")


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Uso: python generate-production-order.py <json_file> <template> <output>")
        sys.exit(1)
    try:
        json_path     = sys.argv[1]
        template_path = sys.argv[2]
        output_path   = sys.argv[3]
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        generar_op(data, template_path, output_path)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)