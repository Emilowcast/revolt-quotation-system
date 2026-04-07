"""
Generador de Orden de Producción - REVOLT SGAC
Uso: python3 generate_orden_produccion.py '<json_data>' output.xlsx
"""

import sys
import json
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def generate_orden(data: dict, template_path: str, output_path: str):
    # Copiar plantilla para preservar imagen, estilos y estructura exacta
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # ── Helpers ──────────────────────────────────────────────────
    def set_val(coord, value):
        """Escribe valor preservando el estilo existente de la celda."""
        cell = ws[coord]
        cell.value = value

    def fmt_date(iso_str):
        if not iso_str:
            return ''
        try:
            return datetime.fromisoformat(iso_str.replace('Z', '+00:00')).strftime('%d/%m/%Y')
        except Exception:
            return iso_str

    # ── Extraer datos del JSON ────────────────────────────────────
    folio        = data.get('quoteFollio') or data.get('folio', '')
    fecha        = fmt_date(data.get('date') or data.get('createdAt', ''))
    _cname   = data.get('clientName', '')
    _cempresa = data.get('clientCompany', '')
    if _cempresa and _cname:
        cliente = f"{_cname} - {_cempresa}"
    elif _cempresa:
        cliente = _cempresa
    else:
        cliente = _cname
    notas        = data.get('notes', '')
    prioridad    = data.get('priority', 'normal').upper()
    fecha_inicio = fmt_date(data.get('createdAt', ''))
    fecha_salida = fmt_date(data.get('deliveryDate') or data.get('dueDate', ''))
    adicionales  = data.get('additionalNotes', '')

    items = data.get('items', [])

    # ── CABECERA ─────────────────────────────────────────────────
    # Labels en N2/P2 se mantienen, valores van en la fila de abajo
    ws['N2'].value = 'FECHA'
    ws['P2'].value = 'FOLIO'
    ws['N3'].value = fecha
    ws['P3'].value = folio

    # CLIENTE → M5 (merged M5:P5) — campo de valor
    set_val('M5', cliente)

    # ── INFORMACIÓN DEL EQUIPO (items) ───────────────────────────
    # Fila 6-7 = ENCABEZADOS, Fila 8 = VALORES del equipo
    if items:
        item = items[0]
        producto    = item.get('modelo', item.get('model', ''))
        descripcion = item.get('descripcion', item.get('description', ''))
        cantidad    = item.get('cant', item.get('qty', 1))

        capacidad       = ''
        voltaje_entrada = ''
        voltaje_salida  = ''
        amp_entrada     = ''
        amp_salida      = ''
        no_serie        = ''
        tipo_fase       = ''

        for line in descripcion.split('\n'):
            line = line.strip()
            ll = line.lower()
            if 'capacidad' in ll or 'kva' in ll or 'kw' in ll:
                capacidad = line.split(':', 1)[-1].strip() if ':' in line else line
            elif 'voltaje entrada' in ll or 'v entrada' in ll:
                voltaje_entrada = line.split(':', 1)[-1].strip() if ':' in line else line
            elif 'voltaje salida' in ll or 'v salida' in ll or ('salida' in ll and 'voltaje' in ll) or ('salida' in ll and 'volt' in ll):
                voltaje_salida = line.split(':', 1)[-1].strip() if ':' in line else line
            elif 'amperaje entrada' in ll or 'amp entrada' in ll or 'a entrada' in ll:
                amp_entrada = line.split(':', 1)[-1].strip() if ':' in line else line
            elif 'amperaje salida' in ll or 'amp salida' in ll or 'a salida' in ll:
                amp_salida = line.split(':', 1)[-1].strip() if ':' in line else line
            elif 'serie' in ll or 'serial' in ll:
                no_serie = line.split(':', 1)[-1].strip() if ':' in line else line
            elif 'monofasico' in ll or '1f' in ll or '1 f' in ll:
                tipo_fase = '1F'
            elif 'bifasico' in ll or '2f' in ll:
                tipo_fase = '2F'
            elif 'trifasico' in ll or '3f' in ll:
                tipo_fase = '3F'

        # PRODUCTO  → A8 (fila de valores, merged A8:B8)
        set_val('A8', producto)

        # CAPACIDAD → C8 (fila de valores)
        set_val('C8', capacidad)

        # VOLTAJE Entrada → E8, Salida → H8
        set_val('E8', voltaje_entrada)
        set_val('H8', voltaje_salida)

        # AMPERAJE Entrada → I8, Salida → L8
        set_val('I8', amp_entrada)
        set_val('L8', amp_salida)

        # MODELO → M8
        set_val('M8', producto)

        # CANTIDAD → N8 (merged N8:P8)
        set_val('N8', str(cantidad))

        # No. SERIE → K10 (merged K10:P10, fila de valores bajo encabezado K9)
        set_val('K10', no_serie)

        # TIPO (1F, 2F, 3F) — marcar con "X" debajo del encabezado correspondiente
        # Encabezados en fila 10: A10=1F, C10=2F, E10=2FN, H10=3F
        # Las "X" van en la misma fila 10 reemplazando el texto del checkbox
        tipo_map = {'1F': 'A10', '2F': 'C10', '2FN': 'E10', '3F': 'H10'}
        if tipo_fase and tipo_fase in tipo_map:
            ws[tipo_map[tipo_fase]].value = ws[tipo_map[tipo_fase]].value + '  ✓'

    # Si hay múltiples ítems, agregar resumen en ADICIONALES
    if len(items) > 1:
        resumen = '\n'.join([
            f"• {it.get('modelo', it.get('model', ''))} x{it.get('cant', it.get('qty', 1))}"
            for it in items[1:]
        ])
        adicionales = (adicionales + '\n' + resumen).strip() if adicionales else resumen

    # ── FECHAS DE PRODUCCIÓN ─────────────────────────────────────
    # Encabezado FECHA INICIO en A12, valor en G12 (merged G12:J13)
    set_val('G12', fecha_inicio)
    # Encabezado FECHA SALIDA en A14, valor en G14 (merged G14:J15)
    set_val('G14', fecha_salida)

    # ADICIONALES → K13 (fila de valor bajo encabezado K12)
    if adicionales:
        set_val('K13', adicionales)

    # ── OBSERVACIONES ────────────────────────────────────────────
    # Encabezado OBSERVACIONES en A16, valores en filas 17-19
    obs_lines = []
    if notas:
        obs_lines.append(notas)
    if prioridad and prioridad != 'NORMAL':
        obs_lines.append(f"PRIORIDAD: {prioridad}")

    if obs_lines:
        set_val('A17', '\n'.join(obs_lines))

    # ── AJUSTE COLUMNA P para que el folio sea visible ──────────
    folio_len = len(str(folio)) + 2
    if folio_len > (ws.column_dimensions['P'].width or 12):
        ws.column_dimensions['P'].width = folio_len

    # ── WRAP_TEXT en celdas con contenido largo ──────────────────
    from openpyxl.styles import Alignment as Align

    wrap_cells = ['A8', 'A17', 'A18', 'A19', 'K13', 'M5', 'N3', 'P3', 'M8']
    for coord in wrap_cells:
        cell = ws[coord]
        if cell.value:
            cell.alignment = Align(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical or 'top',
                wrap_text=True
            )

    # ── GUARDAR ──────────────────────────────────────────────────
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Uso: python3 generate_orden_produccion.py '<json>' <template.xlsx> <output.xlsx>")
        sys.exit(1)

    data = json.loads(sys.argv[1])
    template = sys.argv[2]
    output = sys.argv[3]
    generate_orden(data, template, output)
    print(f"OK:{output}")